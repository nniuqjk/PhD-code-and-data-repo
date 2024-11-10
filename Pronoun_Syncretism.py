#!/usr/bin/env python
#coding=utf-8

"""
Generates a binary strings for paradigms and writes it to a nexus file.

"""

import os
import binascii

try:
    from openpyxl import load_workbook
except ImportError:
    raise ImportError("Please install openpyxl")

try:
    from nexus import NexusWriter
except ImportError:
    raise ImportError("Please install python-nexus")

#open the xl file and read the rows into the sibling term dictionary


def get_data(xlsfile):
    wb = load_workbook(filename=xlsfile)
    w = wb.worksheets[0]
    siblings = []
    header = None
    sibline = {}
    print(w.iter_rows())
    for row in w.iter_rows():
        row = [r.value for r in row]
        if header is None:
            header = row
        else:
            sibline = dict(zip(header, row))
            print(sibline)
            sibline.pop(None,None)
            siblings.append(sibline)
            
    return siblings

# create a dictionary containing canonical representations of the paradigm cells

canon = {'1PduexA' : 0b100010001010000,   #21
         '1PduexS' : 0b010010001010000,   #22
         '1PduexO' : 0b001010001010000,   #25
         '1PduexP' : 0b000110001010000,   #26
         '1PduinA' : 0b100010010010000,   #69
         '1PduinS' : 0b010010010010000,   #70
         '1PduinO' : 0b001010010010000,   #41
         '1PduinP' : 0b000110010010000,   #42
         '1PplexA' : 0b100010001001000,   #21
         '1PplexS' : 0b010010001001000,   #22
         '1PplexO' : 0b001010001001000,   #25
         '1PplexP' : 0b000110001001000,   #26
         '1PplinA' : 0b100010010001000,   #69
         '1PplinS' : 0b010010010001000,   #70
         '1PplinO' : 0b001010010001000,   #41
         '1PplinP' : 0b000110010001000,   #42
         '1PsgA' : 0b100010000100000,   #21
         '1PsgS' : 0b010010000100000,   #22
         '1PsgO' : 0b001010000100000,   #25
         '1PsgP' : 0b000110000100000,   #26
         '2PduA' : 0b100001000010000,   #69
         '2PduS' : 0b010001000010000,   #70
         '2PduO' : 0b001001000010000,   #41
         '2PduP' : 0b000101000010000,   #42
         '2PplA' : 0b100001000001000,   #26
         '2PplS' : 0b010001000001000,   #69
         '2PplO' : 0b001001000001000,   #70
         '2PplP' : 0b000101000001000,   #41
         '2PsgA' : 0b100001000100000,   #42
         '2PsgS' : 0b010001000100000,   #26
         '2PsgO' : 0b001001000100000,   #69
         '2PsgP' : 0b000101000100000,   #70
         '3PduA' : 0b100000100010000,   #41
         '3PduS' : 0b010000100010000,   #42
         '3PduO' : 0b001000100010000,   #26
         '3PduP' : 0b000100100010000,   #69
         '3PplA' : 0b100000100001000,   #70
         '3PplS' : 0b010000100001000,   #41
         '3PplO' : 0b001000100001000,   #42
         '3PplP' : 0b000100100001000,   #26
         '3PsgmA' : 0b100000100100100,   #69
         '3PsgmS' : 0b010000100100100,   #70
         '3PsgmO' : 0b001000100100100,   #41
         '3PsgmP' : 0b000100100100100,   #42
         '3PsgnA' : 0b100000100100010,   #26
         '3PsgnS' : 0b010000100100010,   #69
         '3PsgnO' : 0b001000100100010,   #70
         '3PsgnP' : 0b000100100100010,   #41
         }
# creating the binary data model of the paradigm

def create_model(data):
    data.pop(0)
    model_dict = []
#    i = 0
    for x in data:
        dictline = {}
        for k1, v1 in x.items():
#            print(k1)
            a = "Language"
            b = "ISOCODE"
            c = "qqq"
            if (k1 == a):
                dictline[k1] = v1
                continue
            elif (k1 == b):
                dictline[k1] = v1
                continue
            elif (v1 == c):
              dictline[k1] = "?????????????????????????????????????"
              continue
            equivalence = canon.get(k1)
            #print(equivalence)
            for k2, v2 in x.items():
                #  print(k2)
                  if (k2 == a):
                    continue
                  elif (k2 == b):
                    continue
                  elif (k2 == k1):
                    continue
                  elif (v2 == v1):
                    equivalence = equivalence|canon.get(k2)
                  dictline[k1] = bin(equivalence)
                  print(dictline)
        model_dict.append(dictline)
    return(model_dict)
       
                    

def make_nexus(model_dict):
    sequence = []
    n = NexusWriter()
    dictline = {}
    for dictline in model_dict:
        lang = dictline.get('Language')
        isocode = dictline.get('ISOCODE')
        kvec = ""
        for k in sorted(dictline):
            print(k)
            if (k == 'Language' or k == 'ISOCODE'):
                continue
            else:
                k = repr(dictline.get(k))
                krem = k.replace("'", "")
                kmod = krem[2:16]
                kvec = kvec + kmod.zfill(14)
        label = "%s_%s_%d" % ("48", "113", 1)
        n.add(lang, label, kvec)
    return n

if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Generates a nexus')
    # positional
    parser.add_argument("xlsfile", help="Excel file")
    parser.add_argument("outfile", help="Output file")
    args = parser.parse_args()
    
    if os.path.isfile(args.outfile):
        raise IOError(
            "Output file %s already exists, please rename" % args.outfile
        )
    
    if not os.path.isfile(args.xlsfile):
        raise IOError(
            "Unable to find input xlsx file %s" % args.xlsfile
        )
    
    data = get_data(args.xlsfile)
    model_dict = create_model(data)
    nex = make_nexus(model_dict)
    nex.write_to_file(args.outfile, charblock=True)